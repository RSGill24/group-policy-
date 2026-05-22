"""
Planview2026PlatinumConsolidatedViewCrossDB.py
===============================================
Cross-database view variant of the Planview Prod Data Pipeline.
Views live in a different database (view_database) to the input/output
schemas which stay in the main database (database).

All downstream logic (SP, mappings, output Excel, amber highlighting,
_Original columns, run_history) is identical to the main pipeline.

Main files untouched:
    Planview2026PlatinumConsolidatedUpdated.py       — Excel-based pipeline
    Planview2026PlatinumConsolidatedView.py          — Single-DB view pipeline
    Planview_2026_Platinum_Consolidated_Updated.sql  — stored procedure
"""

import pandas as pd
import pyodbc
import json
import argparse
import sys
import io
import re
import logging
from pathlib import Path
from datetime import datetime

# ── Config globals ─────────────────────────────────────────────
VIEW_INITIATIVES  = "vw_Initiatives"
VIEW_EPICS        = "vw_Epics"
VIEW_SBA          = "vw_SBA"
VIEW_TASKS        = "vw_Tasks"
VIEW_SCHEMA       = "dbo"
VIEW_DATABASE     = ""        # source DB where views live
OUTPUT_FOLDER     = ""
SQL_SERVER        = ""
SQL_DATABASE      = ""        # working DB for input/output schemas
NRB_FIELD         = "L1 Net Recurring Benefits ($, annualized)-P&L/Hard"
NRB_THRESHOLD_M   = 10
_logger           = None
SEPARATOR         = "=" * 60


# ── Logging ───────────────────────────────────────────────────
def init_logger(ts, out_folder):
    global _logger
    Path(out_folder).mkdir(parents=True, exist_ok=True)
    log_file = Path(out_folder) / f"prod_pipeline_run_{ts}.log"
    _logger = logging.getLogger("prod_pipeline")
    _logger.setLevel(logging.DEBUG)
    _logger.handlers.clear()
    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setFormatter(logging.Formatter("%(asctime)s  %(message)s",
                                       datefmt="%Y-%m-%d %H:%M:%S"))
    _logger.addHandler(fh)
    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(logging.Formatter("%(message)s"))
    _logger.addHandler(ch)
    return log_file

def log(msg, indent=0):
    line = "  " * indent + msg
    _logger.info(line) if _logger else print(line)

def log_step(n, msg):
    _logger.info(f"\n[{n}] {msg}") if _logger else print(f"\n[{n}] {msg}")


# ── Config loader ─────────────────────────────────────────────
def load_config(config_path=None):
    global VIEW_INITIATIVES, VIEW_EPICS, VIEW_SBA, VIEW_TASKS
    global VIEW_SCHEMA, VIEW_DATABASE
    global OUTPUT_FOLDER, SQL_SERVER, SQL_DATABASE, NRB_FIELD, NRB_THRESHOLD_M

    if config_path is None:
        config_path = Path(__file__).parent / "Planview2026PlatinumConsolidated_ViewCrossDB_config.json"
    config_path = Path(config_path)
    if not config_path.exists():
        print(f"ERROR: Config not found: {config_path}"); sys.exit(1)
    try:
        with open(config_path, encoding="utf-8") as f:
            cfg = json.load(f)
    except json.JSONDecodeError as e:
        print(f"ERROR: Bad JSON: {e}"); sys.exit(1)

    OUTPUT_FOLDER     = cfg["output_folder"]
    SQL_SERVER        = cfg.get("sql", {}).get("server", "")
    SQL_DATABASE      = cfg.get("sql", {}).get("database", "")
    VIEW_DATABASE     = cfg.get("sql", {}).get("view_database", "")
    VIEW_SCHEMA       = cfg.get("view_schema", "dbo")
    VIEW_INITIATIVES  = cfg.get("view_initiatives", "vw_Initiatives")
    VIEW_EPICS        = cfg.get("view_epics",       "vw_Epics")
    VIEW_SBA          = cfg.get("view_sba",         "vw_SBA")
    VIEW_TASKS        = cfg.get("view_tasks",       "vw_Tasks")
    NRB_FIELD         = cfg.get("nrb_field",
                                "L1 Net Recurring Benefits ($, annualized)-P&L/Hard")
    NRB_THRESHOLD_M   = cfg.get("nrb_threshold_m", 10)


# ── Read from SQL view (cross-database) ──────────────────────
def read_from_view(conn, view_name):
    """
    Reads all rows from a SQL view using a three-part name:
        [view_database].[view_schema].[view_name]
    Falls back to single-part name if view_database is not set.
    """
    if VIEW_DATABASE:
        full_name = f"[{VIEW_DATABASE}].[{VIEW_SCHEMA}].[{view_name}]"
    else:
        full_name = f"[{view_name}]"
    if not view_name or not view_name.strip():
        log(f"  [skipped — view name not configured]", 1)
        return pd.DataFrame()
    try:
        cursor = conn.cursor()
        cursor.execute(f"SELECT * FROM {full_name}")
        cols = [col[0] for col in cursor.description]
        rows = cursor.fetchall()
        df = pd.DataFrame.from_records([tuple(r) for r in rows], columns=cols)
        log(f"  {full_name} — {len(df):,} rows | {len(df.columns)} cols", 1)
        return df
    except Exception as e:
        log(f"ERROR reading view {full_name}: {e}", 1)
        sys.exit(1)


# ── SQL: Connect ──────────────────────────────────────────────
def connect_sql():
    log_step("2/5", "Connecting to SQL Server...")
    try:
        conn = pyodbc.connect(
            f"DRIVER={{ODBC Driver 17 for SQL Server}};"
            f"SERVER={SQL_SERVER};DATABASE={SQL_DATABASE};"
            f"Trusted_Connection=yes;Connection Timeout=30;"
        )
        conn.autocommit = True
        log(f"Server   : {SQL_SERVER} / {SQL_DATABASE}", 1)
        log("Status   : Connected", 1)
        return conn
    except pyodbc.Error as e:
        log(f"ERROR: Could not connect to SQL Server: {e}", 1)
        sys.exit(1)


# ── SQL column name sanitiser ─────────────────────────────────
def _sql_col(name):
    clean = (str(name)
             .replace(']', '')
             .replace('[', '')
             .replace(',', '_')
             .replace('/', '_')
             .replace('\\', '_')
             .replace(':', '_')
             .replace('?', '')
             .replace('(', '')
             .replace(')', '')
             .strip())
    return clean[:120]


# ── SQL: Load single dataframe into a table ───────────────────
def load_to_sql(conn, df, schema_name, table_name, ts):
    cursor = conn.cursor()
    cursor.execute(f"""
        IF NOT EXISTS (SELECT 1 FROM sys.schemas WHERE name='{schema_name}')
            EXEC('CREATE SCHEMA [{schema_name}]')
    """)
    cursor.execute(f"""
        IF OBJECT_ID('[{schema_name}].[{table_name}]') IS NOT NULL
            DROP TABLE [{schema_name}].[{table_name}]
    """)

    safe_cols = [_sql_col(c) for c in df.columns]
    seen = {}
    final_cols = []
    for c in safe_cols:
        if c in seen:
            seen[c] += 1
            c = f"{c}_{seen[c]}"
        else:
            seen[c] = 0
        final_cols.append(c)
    df = df.copy()
    df.columns = final_cols

    # Drop pipeline metadata cols if view was built on a previously saved output table
    df = df.drop(columns=[c for c in ['Run_ID', 'Load_Timestamp', 'Save_Timestamp']
                           if c in df.columns])

    col_defs = ", ".join([f"[{c}] nvarchar(MAX)" for c in df.columns])
    cursor.execute(f"""
        CREATE TABLE [{schema_name}].[{table_name}] (
            [Run_ID]         nvarchar(50)  DEFAULT '{ts}',
            [Load_Timestamp] datetime      DEFAULT GETDATE(),
            {col_defs}
        )
    """)
    col_names    = ", ".join([f"[{c}]" for c in df.columns])
    placeholders = ", ".join(["?" for _ in df.columns])
    ins = (
        f"INSERT INTO [{schema_name}].[{table_name}] "
        f"([Run_ID],[Load_Timestamp],{col_names}) "
        f"VALUES ('{ts}',GETDATE(),{placeholders})"
    )
    df_c = df.copy()
    # Convert ALL columns to string — view may return numeric/float/date types
    # which pyodbc cannot insert into nvarchar(MAX) cleanly.
    # None/NaN/NaT → empty string; everything else → str().strip()
    import math
    def safe_str(v):
        if v is None:
            return ''
        try:
            if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
                return ''
        except Exception:
            pass
        s = str(v).strip()
        return (s
            .replace("\n", " ")
            .replace("\r", " ")
            .replace("|",  " ")
            .replace('"', "'"))

    for col in df_c.columns:
        df_c[col] = df_c[col].apply(safe_str)

    rows = [tuple(r) for r in df_c.itertuples(index=False, name=None)]
    conn.autocommit = False
    try:
        for i in range(0, len(rows), 500):
            cursor.executemany(ins, rows[i:i+500])
        conn.commit()
    except Exception as e:
        conn.rollback(); raise
    finally:
        conn.autocommit = True
    log(f"  [{schema_name}].[{table_name}] — {len(rows):,} rows loaded", 1)


# ── SQL: Load raw input into input schema ─────────────────────
def create_input_schema_and_load(conn, df_inits, df_epics, df_sba, df_tasks, ts):
    log_step("2a/5", "Loading view data to SQL Server input schema...")
    schema   = f"input_{ts}"
    stem     = "Planview_View"
    tbl_init = f"{stem}_Initiatives"
    tbl_epic = f"{stem}_Epics"
    tbl_sba  = f"{stem}_SBA"
    tbl_task = f"{stem}_Tasks"
    load_to_sql(conn, df_inits, schema, tbl_init, ts)
    load_to_sql(conn, df_epics, schema, tbl_epic, ts)
    if not df_sba.empty:
        load_to_sql(conn, df_sba, schema, tbl_sba, ts)
    else:
        log(f"  SBA view not configured — skipping load", 1)
    if not df_tasks.empty:
        load_to_sql(conn, df_tasks, schema, tbl_task, ts)
    else:
        # SP requires the Tasks table to exist — create an empty placeholder
        log(f"  Tasks view not configured — creating empty placeholder table", 1)
        cursor = conn.cursor()
        cursor.execute(f"""
            IF NOT EXISTS (SELECT 1 FROM sys.schemas WHERE name='{schema}')
                EXEC('CREATE SCHEMA [{schema}]')
        """)
        cursor.execute(f"""
            IF OBJECT_ID('[{schema}].[{tbl_task}]') IS NOT NULL
                DROP TABLE [{schema}].[{tbl_task}]
        """)
        cursor.execute(f"""
            CREATE TABLE [{schema}].[{tbl_task}] (
                [_placeholder] nvarchar(10) NULL
            )
        """)
        conn.commit()
    log(f"Input schema : [{schema}]", 1)
    return schema, stem


# ── STEPS 2a–4: Call stored procedure ────────────────────────
def run_transform_sp(conn, ts, input_schema, stem):
    """
    Runs the mapping stored procedure and returns mapped dataframes.

    Returns:
        df_inits     — mapped Initiatives dataframe
        df_epics     — mapped Epics dataframe
        changes_in   — dict {col: cnt}
        changes_ep   — dict {col: cnt}
    """
    log_step("3/5", "Planview2026PlatinumConsolidatedUpdated...")

    nrb_field_sql = _sql_col(NRB_FIELD)

    cursor = conn.cursor()
    cursor.execute(
        "EXEC dbo.Planview2026PlatinumConsolidatedUpdated "
        "    @RunID=?, @InputSchema=?, @Stem=?, @NRB_Field=?, @NRB_Threshold_M=?",
        ts, input_schema, stem, nrb_field_sql, float(NRB_THRESHOLD_M)
    )

    # ── Result set 1: mapped Initiatives ─────────────────────────────────────
    cols_in = [col[0] for col in cursor.description]
    rows_in = cursor.fetchall()
    df_inits = pd.DataFrame.from_records(
        [tuple(r) for r in rows_in], columns=cols_in
    )
    df_inits = df_inits.drop(
        columns=[c for c in ['Run_ID', 'Load_Timestamp'] if c in df_inits.columns]
    )

    # ── Result set 2: mapped Epics ────────────────────────────────────────────
    cursor.nextset()
    cols_ep = [col[0] for col in cursor.description]
    rows_ep = cursor.fetchall()
    df_epics = pd.DataFrame.from_records(
        [tuple(r) for r in rows_ep], columns=cols_ep
    )
    df_epics = df_epics.drop(
        columns=[c for c in ['Run_ID', 'Load_Timestamp'] if c in df_epics.columns]
    )

    # ── Result set 3: scalar counts (all 0) ──────────────────────────────────
    cursor.nextset()
    cursor.fetchone()  # consume RS3

    # ── Result set 4: changes_in ──────────────────────────────────────────────
    cursor.nextset()
    changes_in = {}
    for r in cursor.fetchall():
        if r.cnt and r.cnt > 0:
            changes_in[r.col] = r.cnt

    # ── Result set 5: changes_ep ──────────────────────────────────────────────
    cursor.nextset()
    changes_ep = {}
    for r in cursor.fetchall():
        if r.cnt and r.cnt > 0:
            changes_ep[r.col] = r.cnt

    # ── Result set 6: mapped Tasks ────────────────────────────────────────────
    cursor.nextset()
    cols_tk = [col[0] for col in cursor.description]
    rows_tk = cursor.fetchall()
    df_tasks = pd.DataFrame.from_records(
        [tuple(r) for r in rows_tk], columns=cols_tk
    )
    df_tasks = df_tasks.drop(
        columns=[c for c in ['Run_ID', 'Load_Timestamp', '_placeholder'] if c in df_tasks.columns]
    )

    log(f"Step 2a — Value transformations:", 1)
    if changes_in:
        for col, cnt in changes_in.items():
            log(f"Initiatives — {col}: {cnt:,} values remapped", 2)
    if changes_ep:
        for col, cnt in changes_ep.items():
            log(f"Epics       — {col}: {cnt:,} values remapped", 2)
    if not changes_in and not changes_ep:
        log("No values remapped", 2)

    log(f"Mapping complete:", 1)
    log(f"Initiatives : {len(df_inits):,} rows | {len(df_inits.columns)} cols", 2)
    log(f"Epics       : {len(df_epics):,} rows | {len(df_epics.columns)} cols", 2)
    log(f"Tasks       : {len(df_tasks):,} rows | {len(df_tasks.columns)} cols", 2)

    return df_inits, df_epics, df_tasks, changes_in, changes_ep


# ── SQL: Save single output table ─────────────────────────────
def save_output_to_sql(conn, df, schema_name, table_name, ts):
    cursor = conn.cursor()
    cursor.execute(f"""
        IF NOT EXISTS (SELECT 1 FROM sys.schemas WHERE name='{schema_name}')
            EXEC('CREATE SCHEMA [{schema_name}]')
    """)
    cursor.execute(f"""
        IF OBJECT_ID('[{schema_name}].[{table_name}]') IS NOT NULL
            DROP TABLE [{schema_name}].[{table_name}]
    """)
    if df.empty:
        log(f"  [{schema_name}].[{table_name}] — 0 rows (skipped)", 1)
        return

    safe_cols = [_sql_col(c) for c in df.columns]
    seen = {}
    final_cols = []
    for c in safe_cols:
        if c in seen:
            seen[c] += 1
            c = f"{c}_{seen[c]}"
        else:
            seen[c] = 0
        final_cols.append(c)
    df = df.copy()
    df.columns = final_cols

    # Drop pipeline metadata cols if view was built on a previously saved output table
    df = df.drop(columns=[c for c in ['Run_ID', 'Load_Timestamp', 'Save_Timestamp']
                           if c in df.columns])

    col_defs = ", ".join([f"[{c}] nvarchar(MAX)" for c in df.columns])
    cursor.execute(f"""
        CREATE TABLE [{schema_name}].[{table_name}] (
            [Run_ID]         nvarchar(50) DEFAULT '{ts}',
            [Save_Timestamp] datetime     DEFAULT GETDATE(),
            {col_defs}
        )
    """)
    col_names    = ", ".join([f"[{c}]" for c in df.columns])
    placeholders = ", ".join(["?" for _ in df.columns])
    ins = (
        f"INSERT INTO [{schema_name}].[{table_name}] "
        f"([Run_ID],[Save_Timestamp],{col_names}) "
        f"VALUES ('{ts}',GETDATE(),{placeholders})"
    )
    import math
    def safe_str(v):
        if v is None:
            return ''
        try:
            if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
                return ''
        except Exception:
            pass
        s = str(v).strip()
        return (s
            .replace("\n", " ")
            .replace("\r", " ")
            .replace("|",  " ")
            .replace('"', "'"))

    df_c = df.copy()
    for col in df_c.columns:
        df_c[col] = df_c[col].apply(safe_str)
    rows = [tuple(r) for r in df_c.itertuples(index=False, name=None)]
    conn.autocommit = False
    try:
        for i in range(0, len(rows), 500):
            cursor.executemany(ins, rows[i:i+500])
        conn.commit()
    except Exception as e:
        conn.rollback(); raise
    finally:
        conn.autocommit = True
    log(f"  [{schema_name}].[{table_name}] — {len(df):,} rows saved", 1)


# ── SQL: Save all output sets to output schema ────────────────
def create_output_schema_and_save(conn, df_inits, df_epics, df_sba, df_tasks, stem, ts):
    log_step("5a/5", "Saving mapped output to SQL Server...")
    schema = f"output_{ts}"
    save_output_to_sql(conn, df_inits, schema, f"{stem}_Initiatives", ts)
    save_output_to_sql(conn, df_epics, schema, f"{stem}_Epics",       ts)
    save_output_to_sql(conn, df_sba,   schema, f"{stem}_SBA",         ts)
    save_output_to_sql(conn, df_tasks, schema, f"{stem}_Tasks",       ts)
    log(f"Output schema: [{schema}]", 1)
    return schema

# ── Helper: Add _Original columns next to transformed columns ─
def add_original_cols(df_output, df_raw, mapped_cols, join_key):
    """
    Inserts <col>_Original immediately after each transformed column.
    Uses join_key to match output rows back to the correct input row.
    Handles duplicate join key values by keeping first occurrence.
    Handles cases where _sql_col() strips special chars (e.g. '?') from
    column names — tries the raw name and the raw name + '?' as fallback.
    """
    df = df_output.copy()

    if join_key not in df_raw.columns:
        return df

    # Deduplicate on join_key — keep first occurrence to avoid InvalidIndexError
    df_raw_dedup = df_raw.drop_duplicates(subset=[join_key], keep='first')
    raw_indexed  = df_raw_dedup.set_index(join_key)

    for col in mapped_cols:
        if col not in df.columns:
            continue

        raw_col = None
        if col in df_raw.columns:
            raw_col = col
        elif col + '?' in df_raw.columns:
            raw_col = col + '?'

        if raw_col is None:
            continue

        orig_col = col + '_Original'
        col_idx  = df.columns.get_loc(col)

        orig_vals = df[join_key].map(
            raw_indexed[raw_col]
        ).fillna('').astype(str).str.strip()

        df.insert(col_idx + 1, orig_col, orig_vals.values)

    return df


def get_changed_positions(df_with_orig, mapped_cols):
    """
    Returns set of (row, col) 1-based Excel positions where value changed.
    Compares <col> vs <col>_Original for each mapped column.
    Row 1 = header so data starts at row 2.
    """
    changed = set()
    for col in mapped_cols:
        orig_col = col + '_Original'
        if col not in df_with_orig.columns or orig_col not in df_with_orig.columns:
            continue
        col_i  = df_with_orig.columns.get_loc(col) + 1
        orig_i = df_with_orig.columns.get_loc(orig_col) + 1
        for row_i, (nv, ov) in enumerate(
            zip(df_with_orig[col], df_with_orig[orig_col]), start=2
        ):
            nv_s = str(nv).strip() if nv is not None else ''
            ov_s = str(ov).strip() if ov is not None else ''
            if nv_s != ov_s:
                changed.add((row_i, col_i))
                changed.add((row_i, orig_i))
    return changed


# ── STEP 5b: Build Excel output ───────────────────────────────
def build_excel(df_inits, df_epics, df_sba, df_tasks, df_in_raw, df_ep_raw, df_tk_raw,
                ts, input_path, out_folder, changes_in, changes_ep):

    from openpyxl.styles import PatternFill
    AMBER = PatternFill("solid", start_color="FFE699", fgColor="FFE699")

    def _sanitize_df(df):
        """Strip openpyxl illegal characters from all string cells."""
        try:
            from openpyxl.utils.exceptions import IllegalCharacterError
            from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
        except ImportError:
            import re
            ILLEGAL_CHARACTERS_RE = re.compile(r'[\x00-\x08\x0b-\x0c\x0e-\x1f\uFFFE\uFFFF]')
        df = df.copy()
        for col in df.columns:
            if df[col].dtype == object:
                df[col] = df[col].apply(
                    lambda v: ILLEGAL_CHARACTERS_RE.sub('', str(v))
                    if v is not None else v
                )
        return df

    def write_sheet(writer, df, sheet_name, df_raw=None, mapped_cols=None, join_key=None):
        if df.empty:
            pd.DataFrame({"Note": ["No records in this category"]}).to_excel(
                writer, sheet_name=sheet_name, index=False)
            return
        if df_raw is not None and mapped_cols and join_key:
            df_out  = add_original_cols(df, df_raw, mapped_cols, join_key)
            changed = get_changed_positions(df_out, mapped_cols)
        else:
            df_out  = df
            changed = set()
        # Sanitize illegal characters before writing to Excel
        df_out = _sanitize_df(df_out)
        df_out.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        for (ri, ci) in changed:
            ws.cell(row=ri, column=ci).fill = AMBER
        for col_cells in ws.columns:
            mx = max((len(str(c.value)) for c in col_cells if c.value), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(mx + 4, 60)

    output_file = Path(out_folder) / f"Planview_Prod_Migration_Output_{ts}.xlsx"
    Path(out_folder).mkdir(parents=True, exist_ok=True)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        init_mapped   = ['Estimated Annualized Value Range',
                         'Work Type', 'Home Portfolio', 'Demand SubType',
                         'Impacted Portfolios', 'Demand Domain or Portfolio',
                         'Is this confidential', 'Lifecycle Status']
        epic_mapped   = ['Work Type', 'Work Status', 'Governance Level',
                         'Home Domain/Portfolio', 'Impacted Portfolios',
                         'Demand Domain or Portfolio']
        task_mapped   = ['TASK OR MILESTONE TYPE']
        init_join_key = 'Strategy Seq ID'
        epic_join_key = 'Sequence ID'
        task_join_key = 'SEQUENCE ID'

        write_sheet(writer, df_inits, "Initiatives",
                    df_in_raw, init_mapped, init_join_key)
        write_sheet(writer, df_epics, "Epics",
                    df_ep_raw, epic_mapped, epic_join_key)
        write_sheet(writer, df_tasks, "Tasks",
                    df_tk_raw, task_mapped, task_join_key)
        # SBA — passthrough, no mappings applied, no highlighting
        write_sheet(writer, df_sba, "SBA")

        summary = [
            ["RUN INFORMATION", ""],
            ["Run Timestamp", ts],
            ["Input File",    str(input_path)],
            ["", ""],
            ["VALUE MAPPINGS APPLIED (Initiatives)", ""],
        ]
        for col, cnt in changes_in.items():
            summary.append([f"  {col}", f"{cnt:,} values remapped"])
        if not changes_in:
            summary.append(["  None", ""])

        summary += [["", ""], ["VALUE MAPPINGS APPLIED (Epics)", ""]]
        for col, cnt in changes_ep.items():
            summary.append([f"  {col}", f"{cnt:,} values remapped"])
        if not changes_ep:
            summary.append(["  None", ""])

        summary += [
            ["", ""],
            ["INITIATIVES", f"{len(df_inits):,} total"],
            ["EPICS",       f"{len(df_epics):,} total"],
            ["TASKS",       f"{len(df_tasks):,} total"],
            ["SBA",         f"{len(df_sba):,} total (passthrough — no mappings applied)"],
        ]

        df_sum = pd.DataFrame(summary, columns=["Item", "Value"])
        df_sum.to_excel(writer, sheet_name="Summary", index=False)
        ws_sum = writer.sheets["Summary"]
        ws_sum.column_dimensions['A'].width = 50
        ws_sum.column_dimensions['B'].width = 60

    with open(output_file, 'wb') as f:
        f.write(buf.getvalue())

    return output_file


# ── SQL: Log run to run_history ───────────────────────────────
def log_run_history(conn, ts, input_path, input_schema, output_schema,
                    df_inits, df_epics, df_sba, df_tasks, out_path, elapsed, status):
    cursor = conn.cursor()
    cursor.execute("""
        IF NOT EXISTS (SELECT 1 FROM sys.schemas WHERE name='run_history')
            EXEC('CREATE SCHEMA [run_history]')
    """)
    cursor.execute("""
        IF OBJECT_ID('run_history.Pipeline_Runs_Prod') IS NOT NULL
        BEGIN
            IF NOT EXISTS (
                SELECT 1 FROM sys.columns
                WHERE object_id=OBJECT_ID('run_history.Pipeline_Runs_Prod')
                AND name='Task_Count'
            )
            DROP TABLE run_history.Pipeline_Runs_Prod
        END
    """)
    cursor.execute("""
        IF OBJECT_ID('run_history.Pipeline_Runs_Prod') IS NULL
        CREATE TABLE run_history.Pipeline_Runs_Prod (
            Run_ID             nvarchar(50)  PRIMARY KEY,
            Run_Timestamp      datetime      DEFAULT GETDATE(),
            Pipeline_Name      nvarchar(200),
            Pipeline_Version   nvarchar(20),
            Input_File         nvarchar(500),
            Input_Schema       nvarchar(200),
            Output_Schema      nvarchar(200),
            Init_Count         int,
            Epic_Count         int,
            SBA_Count          int,
            Task_Count         int,
            Output_File_Path   nvarchar(500),
            Runtime_Seconds    decimal(10,1),
            Run_Status         nvarchar(50)
        )
    """)
    cursor.execute("""
        INSERT INTO run_history.Pipeline_Runs_Prod (
            Run_ID, Pipeline_Name, Pipeline_Version, Input_File,
            Input_Schema, Output_Schema,
            Init_Count, Epic_Count, SBA_Count, Task_Count,
            Output_File_Path, Runtime_Seconds, Run_Status
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        ts, 'Prod Data Pipeline', 'prod_v6', str(input_path),
        input_schema, output_schema,
        len(df_inits), len(df_epics), len(df_sba), len(df_tasks),
        str(out_path), elapsed, status
    ))
    conn.commit()
    log(f"Run logged : run_history.Pipeline_Runs_Prod — Run_ID: {ts}", 1)


# ── Main ──────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Planview Prod Data Pipeline — View mode")
    parser.add_argument("--config", default=None)
    args = parser.parse_args()
    load_config(args.config)

    start = datetime.now()
    ts    = start.strftime("%Y%m%d_%H%M%S")
    log_file = init_logger(ts, OUTPUT_FOLDER)

    log(SEPARATOR)
    log("  Planview2026PlatinumConsolidatedViewCrossDB (Cross-DB View mode)")
    log(f"  Views    : {VIEW_DATABASE}.{VIEW_SCHEMA} on {SQL_SERVER}")
    log(f"  Working  : {SQL_DATABASE} on {SQL_SERVER}")
    log(f"  Output   : {OUTPUT_FOLDER}")
    log(f"  Started  : {start.strftime('%Y-%m-%d %H:%M:%S')}")
    log(SEPARATOR)

    # ── Step 1 — Connect to SQL and read views ────────────────
    log_step("1/5", "Connecting to SQL and reading views...")
    conn = connect_sql()

    df_in_raw  = read_from_view(conn, VIEW_INITIATIVES)
    df_ep_raw  = read_from_view(conn, VIEW_EPICS)
    df_sba_raw = read_from_view(conn, VIEW_SBA)
    df_tk_raw  = read_from_view(conn, VIEW_TASKS)

    log(f"Initiatives : {len(df_in_raw):,} rows | {len(df_in_raw.columns)} cols", 1)
    log(f"Epics       : {len(df_ep_raw):,} rows | {len(df_ep_raw.columns)} cols", 1)
    log(f"Tasks       : {len(df_tk_raw):,} rows | {len(df_tk_raw.columns)} cols", 1)
    log(f"SBA         : {len(df_sba_raw):,} rows | {len(df_sba_raw.columns)} cols (passthrough)", 1)

    # ── Step 2 — Load view data to SQL input schema ───────────
    input_schema, stem = create_input_schema_and_load(
        conn, df_in_raw, df_ep_raw, df_sba_raw, df_tk_raw, ts)

    # ── Step 3 — Stored procedure (mappings) ──────────────────
    # SBA bypasses the SP entirely — passed through as-is
    (df_inits, df_epics, df_tasks,
     changes_in, changes_ep) = run_transform_sp(conn, ts, input_schema, stem)

    # ── Console summary ───────────────────────────────────────
    log(f"\nINITIATIVES ({len(df_inits):,} total)", 1)
    log(f"EPICS       ({len(df_epics):,} total)", 1)
    log(f"TASKS       ({len(df_tasks):,} total)", 1)
    log(f"SBA         ({len(df_sba_raw):,} total — passthrough)", 1)

    # ── Step 5a — Save mapped output to SQL ──────────────────
    output_schema = create_output_schema_and_save(
        conn, df_inits, df_epics, df_sba_raw, df_tasks, stem, ts)

    # ── Step 5b — Log run ─────────────────────────────────────
    log_step("5b/5", "Logging run to run_history.Pipeline_Runs_Prod...")

    # ── Step 5c — Write Excel output ─────────────────────────
    log_step("5c/5", "Writing output Excel...")
    out_path = build_excel(
        df_inits, df_epics, df_sba_raw, df_tasks,
        df_in_raw, df_ep_raw, df_tk_raw,
        ts, Path(f"view_{ts}"), OUTPUT_FOLDER, changes_in, changes_ep)

    elapsed = round((datetime.now() - start).total_seconds(), 1)

    log(f"\n{SEPARATOR}")
    log("  PIPELINE COMPLETE (Cross-DB View mode)")
    log(f"  Output : {out_path}")
    log(f"  Log    : {log_file}")
    log(f"")
    log(f"  INITIATIVES : {len(df_inits):,} total")
    log(f"  EPICS       : {len(df_epics):,} total")
    log(f"  TASKS       : {len(df_tasks):,} total")
    log(f"  SBA         : {len(df_sba_raw):,} total (passthrough)")
    log(f"  Runtime     : {elapsed}s")

    log_run_history(conn, ts,
                    Path(f"views:{VIEW_DATABASE}.{VIEW_SCHEMA}:{VIEW_INITIATIVES},{VIEW_EPICS},{VIEW_TASKS},{VIEW_SBA}"),
                    input_schema, output_schema,
                    df_inits, df_epics, df_sba_raw, df_tasks,
                    out_path, elapsed, "Completed")

    conn.close()
    log("SQL      : Connection closed", 1)
    log(SEPARATOR)


if __name__ == "__main__":
    main()
