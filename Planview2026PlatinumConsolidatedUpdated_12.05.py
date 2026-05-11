"""
Planview2026PlatinumConsolidatedUpdated.py
===============================
"""

import pandas as pd
import pyodbc
import json
import argparse
import sys
import io
import re
import logging
from pathlib import Path
from datetime import datetime

# ── Config globals ─────────────────────────────────────────────
INPUT_FILE        = ""
INPUT_SHEET_INITS = "Initiatives Extract_05.01"
INPUT_SHEET_EPICS = "Epics Extract_05.01"
OUTPUT_FOLDER     = ""
SQL_SERVER        = ""
SQL_DATABASE      = ""
NRB_FIELD         = "L1 Net Recurring Benefits ($, annualized)-P&L/Hard"
NRB_THRESHOLD_M   = 10
_logger           = None
SEPARATOR         = "=" * 60

# ── Output segment constants ───────────────────────────────────
SEG = {
    "init_bc":        "TE-NonDisc-BusinessContinuity",
    "init_stopwork":  "TE-Disc-StopWork-HOLD",
    "init_strat_inv": "TE-Disc-TransformationalInvestment",
    "init_disc_other":"TE-Disc-Other",
    "init_bo":        "TE-BusinessDemandMgmt",
    "init_pc":        "PC-StrategicProgram-PilotSandbox",
    "init_pending":   "PENDING-Stage-Lifecycle-Undefined",
    "lcm":            "LCM-NonDisc-RunTheBusiness",
    "le_strat":       "LE-Disc-TransformationalInvestment",
    "le_stopwork":    "LE-Disc-StopWork-HOLD",
    "le_disc":        "LE-Disc-Other",
    "bwt_epic":       "BWT-Epic-BelowPPL",
    "milestone":      "MILESTONE-RISK",
    "epic_pending":   "PENDING-Stage-Lifecycle-Undefined",
    "review":         "REVIEW-NoRuleMatched",
}

FLOW = {
    "init_bc":        "Non-Discretionary \u2013 Business Continuity",
    "init_stopwork":  "STOP WORK Discretionary \u2013 Transformational Investment",
    "init_strat_inv": "Discretionary \u2013 Transformational Investment",
    "init_disc_other":"Discretionary \u2013 Other",
    "init_bo":        "Business Demand Management",
    "init_pc":        "Strategic Program (Business + Tech) \u2014 Pilot Sandbox",
    "lcm":            "Non-Discretionary - Run the Business",
    "le_strat":       "Discretionary \u2013 Transformational Investment",
    "le_stopwork":    "STOP WORK Discretionary \u2013 Transformational Investment",
    "le_disc":        "Discretionary \u2013 Other",
    "bwt_epic":       "Business w/Tech Epic \u2014 Below PPL Task",
    "milestone":      "Milestone/Risk \u2014 separate tab (not in Epic migration scope)",
}


# ── Logging ───────────────────────────────────────────────────
def init_logger(ts, out_folder):
    global _logger
    Path(out_folder).mkdir(parents=True, exist_ok=True)
    log_file = Path(out_folder) / f"prod_pipeline_run_{ts}.log"
    _logger = logging.getLogger("prod_pipeline")
    _logger.setLevel(logging.DEBUG)
    _logger.handlers.clear()
    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setFormatter(logging.Formatter("%(asctime)s  %(message)s",
                                       datefmt="%Y-%m-%d %H:%M:%S"))
    _logger.addHandler(fh)
    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(logging.Formatter("%(message)s"))
    _logger.addHandler(ch)
    return log_file

def log(msg, indent=0):
    line = "  " * indent + msg
    _logger.info(line) if _logger else print(line)

def log_step(n, msg):
    _logger.info(f"\n[{n}] {msg}") if _logger else print(f"\n[{n}] {msg}")


# ── Config loader ─────────────────────────────────────────────
def load_config(config_path=None):
    global INPUT_FILE, INPUT_SHEET_INITS, INPUT_SHEET_EPICS
    global OUTPUT_FOLDER, SQL_SERVER, SQL_DATABASE, NRB_FIELD, NRB_THRESHOLD_M

    if config_path is None:
        config_path = Path(__file__).parent / "Planview2026PlatinumConsolidated_Updated_config.json"
    config_path = Path(config_path)
    if not config_path.exists():
        print(f"ERROR: Config not found: {config_path}"); sys.exit(1)
    try:
        with open(config_path, encoding="utf-8") as f:
            cfg = json.load(f)
    except json.JSONDecodeError as e:
        print(f"ERROR: Bad JSON: {e}"); sys.exit(1)

    INPUT_FILE        = cfg["input_file"]
    OUTPUT_FOLDER     = cfg["output_folder"]
    SQL_SERVER        = cfg.get("sql", {}).get("server", "")
    SQL_DATABASE      = cfg.get("sql", {}).get("database", "")
    INPUT_SHEET_INITS = cfg.get("sheet_initiatives", "Initiatives Extract_05.01")
    INPUT_SHEET_EPICS = cfg.get("sheet_epics",        "Epics Extract_05.01")
    NRB_FIELD         = cfg.get("nrb_field",
                                "L1 Net Recurring Benefits ($, annualized)-P&L/Hard")
    NRB_THRESHOLD_M   = cfg.get("nrb_threshold_m", 10)


# ── STEP 1: Read 3-row header input file ─────────────────────
def read_3row_header(path, sheet_name):
    """
    Row 1 = old system field IDs | Row 2 = new system IDs | Row 3 = display names
    Data starts at Row 4. Display names (Row 3) used as column headers.
    """
    # Use openpyxl directly to preserve literal 'None' cell values
    # pandas read_excel treats 'None' as Python None even with dtype=str
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(['' if v is None else str(v).strip() for v in row])
    wb.close()
    import numpy as np
    raw = pd.DataFrame(data)
    row1_ids  = [str(raw.iloc[0, c]).strip() for c in range(raw.shape[1])]
    row2_ids  = [str(raw.iloc[1, c]).strip() for c in range(raw.shape[1])]
    row3_names= [str(raw.iloc[2, c]).strip() for c in range(raw.shape[1])]

    seen = {}
    final_cols = []
    for name in row3_names:
        if not name or name == 'nan':
            name = f"_col_{len(final_cols)}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        final_cols.append(name)

    df_data = raw.iloc[3:].copy()
    df_data.columns = final_cols

    filter_col = next((c for c in final_cols if not c.startswith('_col')), final_cols[0])
    df_data = df_data[df_data[filter_col].str.strip() != ''].reset_index(drop=True)

    return df_data, row1_ids, row2_ids


# ── SQL: Connect ──────────────────────────────────────────────
def connect_sql():
    log_step("2/5", "Connecting to SQL Server...")
    try:
        conn = pyodbc.connect(
            f"DRIVER={{ODBC Driver 17 for SQL Server}};"
            f"SERVER={SQL_SERVER};DATABASE={SQL_DATABASE};"
            f"Trusted_Connection=yes;Connection Timeout=30;"
        )
        conn.autocommit = True
        log(f"Server   : {SQL_SERVER} / {SQL_DATABASE}", 1)
        log("Status   : Connected", 1)
        return conn
    except pyodbc.Error as e:
        log(f"ERROR: Could not connect to SQL Server: {e}", 1)
        sys.exit(1)


# ── SQL column name sanitiser ─────────────────────────────────
def _sql_col(name):
    clean = (str(name)
             .replace(']', '')
             .replace('[', '')
             .replace(',', '_')
             .replace('/', '_')
             .replace('\\', '_')
             .replace(':', '_')
             .replace('?', '')
             .replace('(', '')
             .replace(')', '')
             .strip())
    return clean[:120]


# ── SQL: Load single dataframe into a table ───────────────────
def load_to_sql(conn, df, schema_name, table_name, ts):
    cursor = conn.cursor()
    cursor.execute(f"""
        IF NOT EXISTS (SELECT 1 FROM sys.schemas WHERE name='{schema_name}')
            EXEC('CREATE SCHEMA [{schema_name}]')
    """)
    cursor.execute(f"""
        IF OBJECT_ID('[{schema_name}].[{table_name}]') IS NOT NULL
            DROP TABLE [{schema_name}].[{table_name}]
    """)

    safe_cols = [_sql_col(c) for c in df.columns]
    seen = {}
    final_cols = []
    for c in safe_cols:
        if c in seen:
            seen[c] += 1
            c = f"{c}_{seen[c]}"
        else:
            seen[c] = 0
        final_cols.append(c)
    df = df.copy()
    df.columns = final_cols

    col_defs = ", ".join([f"[{c}] nvarchar(MAX)" for c in df.columns])
    cursor.execute(f"""
        CREATE TABLE [{schema_name}].[{table_name}] (
            [Run_ID]         nvarchar(50)  DEFAULT '{ts}',
            [Load_Timestamp] datetime      DEFAULT GETDATE(),
            {col_defs}
        )
    """)
    col_names    = ", ".join([f"[{c}]" for c in df.columns])
    placeholders = ", ".join(["?" for _ in df.columns])
    ins = (
        f"INSERT INTO [{schema_name}].[{table_name}] "
        f"([Run_ID],[Load_Timestamp],{col_names}) "
        f"VALUES ('{ts}',GETDATE(),{placeholders})"
    )
    # Fill NaN with '' not None — all columns are nvarchar(MAX) so empty string is correct.
    # Using None causes pyodbc to send SQL NULL, and also means the string 'None'
    # that genuinely exists in the data gets lost during itertuples/pyodbc conversion.
    df_c = df.fillna('')
    for col in df_c.columns:
        if df_c[col].dtype == object and hasattr(df_c[col], "str"):
            df_c[col] = (df_c[col]
                .str.replace("\n", " ", regex=False)
                .str.replace("\r", " ", regex=False)
                .str.replace("|",  " ", regex=False)
                .str.replace('"', "'", regex=False)
                .str.strip())
    rows = [tuple(r) for r in df_c.itertuples(index=False, name=None)]
    conn.autocommit = False
    try:
        for i in range(0, len(rows), 500):
            cursor.executemany(ins, rows[i:i+500])
        conn.commit()
    except Exception as e:
        conn.rollback(); raise
    finally:
        conn.autocommit = True
    log(f"  [{schema_name}].[{table_name}] — {len(rows):,} rows loaded", 1)


# ── SQL: Load raw input into input schema ─────────────────────
def create_input_schema_and_load(conn, df_inits, df_epics, input_path, ts):
    log_step("2a/5", "Loading raw input data to SQL Server...")
    schema   = f"input_{ts}"
    stem     = re.sub(r'[^A-Za-z0-9_]', '_', input_path.stem).strip('_')
    tbl_init = f"{stem}_Initiatives"
    tbl_epic = f"{stem}_Epics"
    load_to_sql(conn, df_inits, schema, tbl_init, ts)
    load_to_sql(conn, df_epics, schema, tbl_epic, ts)
    log(f"Input schema : [{schema}]", 1)
    return schema, tbl_init, tbl_epic


# ── STEPS 2a–4c: Call stored procedure ───────────────────────
def run_transform_sp(conn, ts, input_schema, stem):
    """
    Returns:
        df_inits     — fully classified Initiatives dataframe
        df_epics     — fully classified Epics dataframe
        removed_in   — int
        removed_ep   — int
        new_id_count — int
        changes_in   — dict {col: cnt}
        changes_ep   — dict {col: cnt}
    """
    log_step("3/5", "Calling stored procedure Planview2026PlatinumConsolidatedUpdated...")


    # Sanitise NRB field name the same way load_to_sql does
    # so the SP can reference it as a column name inside the loaded table
    nrb_field_sql = _sql_col(NRB_FIELD)

    cursor = conn.cursor()
    cursor.execute(
        "EXEC dbo.Planview2026PlatinumConsolidatedUpdated "
        "    @RunID=?, @InputSchema=?, @Stem=?, @NRB_Field=?, @NRB_Threshold_M=?",
        ts, input_schema, stem, nrb_field_sql, float(NRB_THRESHOLD_M)
    )

    # ── Result set 1: classified Initiatives ─────────────────
    cols_in = [col[0] for col in cursor.description]
    rows_in = cursor.fetchall()
    df_inits = pd.DataFrame.from_records(
        [tuple(r) for r in rows_in], columns=cols_in
    )
    # Drop internal load columns added during raw load — not needed in output
    df_inits = df_inits.drop(
        columns=[c for c in ['Run_ID', 'Load_Timestamp'] if c in df_inits.columns]
    )

    # ── Result set 2: classified Epics ───────────────────────
    cursor.nextset()
    cols_ep = [col[0] for col in cursor.description]
    rows_ep = cursor.fetchall()
    df_epics = pd.DataFrame.from_records(
        [tuple(r) for r in rows_ep], columns=cols_ep
    )
    df_epics = df_epics.drop(
        columns=[c for c in ['Run_ID', 'Load_Timestamp'] if c in df_epics.columns]
    )

    # ── Result set 3: scalar counts ──────────────────────────
    cursor.nextset()
    row = cursor.fetchone()
    removed_in   = int(row.removed_in)
    removed_ep   = int(row.removed_ep)
    new_id_count = int(row.new_id_count)

    # ── Result set 4: changes_in ─────────────────────────────
    cursor.nextset()
    changes_in = {}
    for r in cursor.fetchall():
        if r.cnt and r.cnt > 0:
            changes_in[r.col] = r.cnt

    # ── Result set 5: changes_ep ─────────────────────────────
    cursor.nextset()
    changes_ep = {}
    for r in cursor.fetchall():
        if r.cnt and r.cnt > 0:
            changes_ep[r.col] = r.cnt

    # ── Result set 6: deleted Initiative rows ────────────────
    df_deleted = pd.DataFrame()
    if cursor.nextset():
        cols_del = [col[0] for col in cursor.description]
        rows_del = cursor.fetchall()
        if rows_del:
            df_deleted = pd.DataFrame.from_records(
                [tuple(r) for r in rows_del], columns=cols_del
            )
            df_deleted = df_deleted.drop(
                columns=[c for c in ['Run_ID', 'Load_Timestamp'] if c in df_deleted.columns]
            )

    # Log what came back
    log(f"Step 2a — Value transformations:", 1)
    if changes_in:
        for col, cnt in changes_in.items():
            log(f"Initiatives — {col}: {cnt:,} values remapped", 2)
    if changes_ep:
        for col, cnt in changes_ep.items():
            log(f"Epics       — {col}: {cnt:,} values remapped", 2)
    if not changes_in and not changes_ep:
        log("No values remapped", 2)

    log(f"Step 2a — Deleted (Stage A: L0 / B: SL1): {removed_in} Initiative rows → Deleted_Records tab", 1)
    log(f"Step 4  — Classification complete:", 1)
    log(f"Initiatives : {len(df_inits):,} rows | {len(df_inits.columns)} cols", 2)
    log(f"Epics       : {len(df_epics):,} rows | {len(df_epics.columns)} cols", 2)
    log(f"NewID_Temp  : {new_id_count:,} Epic records assigned", 2)

    return df_inits, df_epics, df_deleted, removed_in, removed_ep, new_id_count, changes_in, changes_ep


# ── SQL: Save single output table — UNCHANGED from original ──
def save_output_to_sql(conn, df, schema_name, table_name, ts):
    cursor = conn.cursor()
    cursor.execute(f"""
        IF NOT EXISTS (SELECT 1 FROM sys.schemas WHERE name='{schema_name}')
            EXEC('CREATE SCHEMA [{schema_name}]')
    """)
    cursor.execute(f"""
        IF OBJECT_ID('[{schema_name}].[{table_name}]') IS NOT NULL
            DROP TABLE [{schema_name}].[{table_name}]
    """)
    if df.empty:
        log(f"  [{schema_name}].[{table_name}] — 0 rows (skipped)", 1)
        return

    safe_cols = [_sql_col(c) for c in df.columns]
    seen = {}
    final_cols = []
    for c in safe_cols:
        if c in seen:
            seen[c] += 1
            c = f"{c}_{seen[c]}"
        else:
            seen[c] = 0
        final_cols.append(c)
    df = df.copy()
    df.columns = final_cols
    col_defs = ", ".join([f"[{c}] nvarchar(MAX)" for c in df.columns])
    cursor.execute(f"""
        CREATE TABLE [{schema_name}].[{table_name}] (
            [Run_ID]         nvarchar(50) DEFAULT '{ts}',
            [Save_Timestamp] datetime     DEFAULT GETDATE(),
            {col_defs}
        )
    """)
    col_names    = ", ".join([f"[{c}]" for c in df.columns])
    placeholders = ", ".join(["?" for _ in df.columns])
    ins = (
        f"INSERT INTO [{schema_name}].[{table_name}] "
        f"([Run_ID],[Save_Timestamp],{col_names}) "
        f"VALUES ('{ts}',GETDATE(),{placeholders})"
    )
    df_c = df.where(df.notna(), None)
    rows = [tuple(r) for r in df_c.itertuples(index=False, name=None)]
    conn.autocommit = False
    try:
        for i in range(0, len(rows), 500):
            cursor.executemany(ins, rows[i:i+500])
        conn.commit()
    except Exception as e:
        conn.rollback(); raise
    finally:
        conn.autocommit = True
    log(f"  [{schema_name}].[{table_name}] — {len(df):,} rows saved", 1)


# ── SQL: Save all output sets to output schema — UNCHANGED ───
def create_output_schema_and_save(conn, df_inits, df_epics, df_deleted, stem, ts):
    log_step("5a/5", "Saving classified output to SQL Server...")
    schema = f"output_{ts}"

    FINAL_I = {SEG['init_bc'], SEG['init_strat_inv'],
               SEG['init_disc_other'], SEG['init_bo']}
    HOLD_I  = {SEG['init_stopwork'], SEG['init_pc']}
    PEND_I  = {SEG['init_pending'], SEG['review']}
    FINAL_E = {SEG['lcm'], SEG['le_strat'], SEG['le_disc'], SEG['bwt_epic']}
    HOLD_E  = {SEG['le_stopwork']}
    PEND_E  = {SEG['epic_pending'], SEG['review']}
    MILE_E  = {SEG['milestone']}

    init_final     = df_inits[df_inits['Output_Segment'].isin(FINAL_I)].copy()
    init_hold      = df_inits[df_inits['Output_Segment'].isin(HOLD_I)].copy()
    init_review    = df_inits[df_inits['Output_Segment'].isin(PEND_I)].copy()
    epic_final     = df_epics[df_epics['Output_Segment'].isin(FINAL_E)].copy()
    epic_hold      = df_epics[df_epics['Output_Segment'].isin(HOLD_E)].copy()
    epic_review    = df_epics[df_epics['Output_Segment'].isin(PEND_E)].copy()
    epic_milestone = df_epics[df_epics['Output_Segment'].isin(MILE_E)].copy()

    save_output_to_sql(conn, init_final,     schema, f"{stem}_Initiatives_Final",     ts)
    save_output_to_sql(conn, init_hold,      schema, f"{stem}_Initiatives_Hold",       ts)
    save_output_to_sql(conn, init_review,    schema, f"{stem}_Initiatives_Review",     ts)
    save_output_to_sql(conn, epic_final,     schema, f"{stem}_Epics_Final",            ts)
    save_output_to_sql(conn, epic_hold,      schema, f"{stem}_Epics_Hold",             ts)
    save_output_to_sql(conn, epic_review,    schema, f"{stem}_Epics_Review",           ts)
    save_output_to_sql(conn, epic_milestone, schema, f"{stem}_Epics_Milestone_Risk",   ts)
    save_output_to_sql(conn, df_deleted,    schema, f"{stem}_Deleted_Records",          ts)

    log(f"Output schema: [{schema}]", 1)
    return (schema,
            init_final, init_hold, init_review,
            epic_final, epic_hold, epic_review, epic_milestone)


# ── Helper: Add _Original columns next to transformed columns ─
def add_original_cols(df_output, df_raw, mapped_cols, join_key):
    """
    Inserts <col>_Original immediately after each transformed column.
    Uses join_key to match output rows back to the correct input row
    so deleted/reordered rows are handled correctly.
    """
    df = df_output.copy()

    # Build lookup: join_key value → original column value
    if join_key not in df_raw.columns:
        return df  # can't join — return unchanged

    raw_indexed = df_raw.set_index(join_key)

    for col in mapped_cols:
        if col not in df.columns or col not in df_raw.columns:
            continue
        orig_col = col + '_Original'
        col_idx  = df.columns.get_loc(col)

        # Map each output row's join key to the original input value
        orig_vals = df[join_key].map(
            raw_indexed[col]
        ).fillna('').astype(str).str.strip()

        df.insert(col_idx + 1, orig_col, orig_vals.values)

    return df


def get_changed_positions(df_with_orig, mapped_cols):
    """
    Returns set of (row, col) 1-based Excel positions where value changed.
    Compares <col> vs <col>_Original for each mapped column.
    Row 1 = header so data starts at row 2.
    """
    changed = set()
    for col in mapped_cols:
        orig_col = col + '_Original'
        if col not in df_with_orig.columns or orig_col not in df_with_orig.columns:
            continue
        col_i  = df_with_orig.columns.get_loc(col) + 1
        orig_i = df_with_orig.columns.get_loc(orig_col) + 1
        for row_i, (nv, ov) in enumerate(
            zip(df_with_orig[col], df_with_orig[orig_col]), start=2
        ):
            nv_s = str(nv).strip() if nv is not None else ''
            ov_s = str(ov).strip() if ov is not None else ''
            if nv_s != ov_s:
                changed.add((row_i, col_i))
                changed.add((row_i, orig_i))
    return changed


# ── STEP 5b: Build Excel output ───────────────────────────────
def build_excel(df_inits, df_epics, df_deleted, df_in_raw, df_ep_raw,
                ts, input_path, out_folder,
                removed_in, removed_ep, changes_in, changes_ep):

    FINAL_I = {SEG['init_bc'], SEG['init_strat_inv'],
               SEG['init_disc_other'], SEG['init_bo']}
    HOLD_I  = {SEG['init_stopwork'], SEG['init_pc']}
    PEND_I  = {SEG['init_pending'], SEG['review']}
    FINAL_E = {SEG['lcm'], SEG['le_strat'], SEG['le_disc'], SEG['bwt_epic']}
    HOLD_E  = {SEG['le_stopwork']}
    PEND_E  = {SEG['epic_pending'], SEG['review']}
    MILE_E  = {SEG['milestone']}

    init_final    = df_inits[df_inits['Output_Segment'].isin(FINAL_I)].copy()
    init_hold     = df_inits[df_inits['Output_Segment'].isin(HOLD_I)].copy()
    init_review   = df_inits[df_inits['Output_Segment'].isin(PEND_I)].copy()
    epic_final    = df_epics[df_epics['Output_Segment'].isin(FINAL_E)].copy()
    epic_hold     = df_epics[df_epics['Output_Segment'].isin(HOLD_E)].copy()
    epic_review   = df_epics[df_epics['Output_Segment'].isin(PEND_E)].copy()
    epic_milestone= df_epics[df_epics['Output_Segment'].isin(MILE_E)].copy()

    from openpyxl.styles import PatternFill
    AMBER = PatternFill("solid", start_color="FFE699", fgColor="FFE699")

    def write_sheet(writer, df, sheet_name, df_raw=None, mapped_cols=None, join_key=None):
        if df.empty:
            pd.DataFrame({"Note": ["No records in this category"]}).to_excel(
                writer, sheet_name=sheet_name, index=False)
            return
        if df_raw is not None and mapped_cols and join_key:
            df_out  = add_original_cols(df, df_raw, mapped_cols, join_key)
            changed = get_changed_positions(df_out, mapped_cols)
        else:
            df_out  = df
            changed = set()
        df_out.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        for (ri, ci) in changed:
            ws.cell(row=ri, column=ci).fill = AMBER
        for col_cells in ws.columns:
            mx = max((len(str(c.value)) for c in col_cells if c.value), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(mx + 4, 60)

    output_file = Path(out_folder) / f"Planview_Prod_Migration_Output_{ts}.xlsx"
    Path(out_folder).mkdir(parents=True, exist_ok=True)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        init_mapped  = ['Stage', 'Demand Type', 'Estimated Annualized Value Range',
                         'Home Portfolio', 'Demand SubType']
        epic_mapped   = ['Work Status', 'Demand Type', 'Home Domain/Portfolio']
        init_join_key = 'Strategy Seq ID'
        epic_join_key = 'Sequence ID'

        write_sheet(writer, init_final,     "Initiatives_Final",
                    df_in_raw, init_mapped, init_join_key)
        write_sheet(writer, init_hold,      "Initiatives_Hold",
                    df_in_raw, init_mapped, init_join_key)
        write_sheet(writer, init_review,    "Initiatives_Review",
                    df_in_raw, init_mapped, init_join_key)
        write_sheet(writer, epic_final,     "Epics_Final",
                    df_ep_raw, epic_mapped, epic_join_key)
        write_sheet(writer, epic_hold,      "Epics_Hold",
                    df_ep_raw, epic_mapped, epic_join_key)
        write_sheet(writer, epic_review,    "Epics_Review",
                    df_ep_raw, epic_mapped, epic_join_key)
        write_sheet(writer, epic_milestone, "Epics_Milestone_Risk",
                    df_ep_raw, epic_mapped, epic_join_key)
        write_sheet(writer, df_deleted,     "Deleted_Records")

        # Summary tab
        summary = [
            ["RUN INFORMATION", ""],
            ["Run Timestamp",  ts],
            ["Input File",     str(input_path)],
            ["", ""],
            ["PRE-PROCESSING", ""],
            ["Deleted rows (Stage A: L0 or B: SL1)", removed_in],
            ["Deleted rows (Epics)",                 removed_ep],
            ["Value mapping applied (Initiatives)",
             f"Work Status: {changes_in.get('Work Status',0)} values remapped"
             if changes_in else "None"],
            ["Value mapping applied (Epics)",
             f"Work Status: {changes_ep.get('Work Status',0)} values remapped"
             if changes_ep else "None"],
            ["", ""],
            ["INITIATIVES", f"{len(df_inits):,} total"],
            ["Final Output",   f"{len(init_final):,}"],
            ["Stop Work HOLD", f"{len(init_hold):,}  (inc. Purple Chip/DRIVE \u2014 Pilot Sandbox)"],
            ["Review/Pending", f"{len(init_review):,}  (SL/L5 stages undefined in rules file)"],
            ["", ""],
            ["EPICS", f"{len(df_epics):,} total"],
            ["Final Output",   f"{len(epic_final):,}  (incl. BwT Epics with NewID_Temp)"],
            ["Stop Work HOLD", f"{len(epic_hold):,}"],
            ["Review/Pending", f"{len(epic_review):,}  (SL stages undefined)"],
            ["Milestone/Risk (separate tab)", f"{len(epic_milestone):,}  "
             "(Work Type = Initiative Milestones & Risks \u2014 not in Epic migration scope)"],
            ["", ""],
            ["NEWID TEMPORARY INDEX", ""],
            ["Purpose", "Epics classified as below-PPL tasks have a temporary NewID-xxxx. "
                        "Links task to parent Work ID. Replaced after import into Planview."],
            ["Records with NewID_Temp",
             f"{(df_epics['NewID_Temp'] != '').sum():,}"
             if 'NewID_Temp' in df_epics.columns else "N/A"],
            ["", ""],
            ["OPEN ITEMS", ""],
            ["BC field (BR_TE_004)",
             "Is this request vital to business continuity? field does NOT exist in this "
             "prod extract. BC classification returns 0 records. Customer to confirm "
             "which field or value identifies BC initiatives in the full prod extract."],
            ["Purple Chip / PC (BR_PC_001)",
             "Purple Chip field does NOT exist in this prod extract. PC/DRIVE Strategic "
             "Program classification returns 0 records. No PC path to Pilot Sandbox in output. "
             "Customer to confirm how PC initiatives are identified in prod extract."],
            ["Epic Stop Work HOLD = 0",
             "Estimated Annualized Value Range field does not exist in Epic sheet of prod "
             "extract (it exists in Initiative sheet only). BR_TE_LE_003 Stop Work proxy "
             "cannot fire. All LE L/XL Epics go to Transformational Investment. Correct "
             "behaviour for this extract."],
            ["Status exclusion",
             "New rules file has no explicit status exclusion rule. All statuses pass through. "
             "confirm: which Lifecycle Status / Work Status values should be excluded?"],
            ["NRB field",
             "Pending Finance confirmation \u2014 using L1 NRB Hard as assumption. "
             "New rules file BR_TE_002 notes: 'What field are we using for NRB?'"],
            ["Stop Work meaning",
             "What does Stop Work mean operationally? Exclude or migrate as cancelled?"],
            ["SL/L5 stages",
             "Target lifecycle step not defined in rules file for SL/L5 stages. "
             f"{len(init_review):,} Initiative records in Review."],
        ]

        df_sum = pd.DataFrame(summary, columns=["Item", "Value"])
        df_sum.to_excel(writer, sheet_name="Summary", index=False)
        ws_sum = writer.sheets["Summary"]
        ws_sum.column_dimensions['A'].width = 40
        ws_sum.column_dimensions['B'].width = 80

    with open(output_file, 'wb') as f:
        f.write(buf.getvalue())

    return (output_file,
            init_final, init_hold, init_review,
            epic_final, epic_hold, epic_review, epic_milestone)


# ── SQL: Log run to run_history — UNCHANGED from original ─────
def log_run_history(conn, ts, input_path, input_schema, output_schema,
                    init_final, init_hold, init_review,
                    epic_final, epic_hold, epic_review, epic_milestone,
                    out_path, elapsed, status, removed_in, removed_ep):
    cursor = conn.cursor()
    cursor.execute("""
        IF NOT EXISTS (SELECT 1 FROM sys.schemas WHERE name='run_history')
            EXEC('CREATE SCHEMA [run_history]')
    """)
    cursor.execute("""
        IF OBJECT_ID('run_history.Pipeline_Runs_Prod') IS NOT NULL
        BEGIN
            IF NOT EXISTS (
                SELECT 1 FROM sys.columns
                WHERE object_id=OBJECT_ID('run_history.Pipeline_Runs_Prod')
                AND name='Removed_Rows'
            )
            DROP TABLE run_history.Pipeline_Runs_Prod
        END
    """)
    cursor.execute("""
        IF OBJECT_ID('run_history.Pipeline_Runs_Prod') IS NULL
        CREATE TABLE run_history.Pipeline_Runs_Prod (
            Run_ID             nvarchar(50)  PRIMARY KEY,
            Run_Timestamp      datetime      DEFAULT GETDATE(),
            Pipeline_Name      nvarchar(200),
            Pipeline_Version   nvarchar(20),
            Input_File         nvarchar(500),
            Input_Schema       nvarchar(200),
            Output_Schema      nvarchar(200),
            Init_Final         int, Init_Hold    int, Init_Review  int,
            Epic_Final         int, Epic_Hold    int, Epic_Review  int,
            Epic_Milestone     int,
            Removed_Rows       int,
            Output_File_Path   nvarchar(500),
            Runtime_Seconds    decimal(10,1),
            Run_Status         nvarchar(50)
        )
    """)
    cursor.execute("""
        INSERT INTO run_history.Pipeline_Runs_Prod (
            Run_ID, Pipeline_Name, Pipeline_Version, Input_File,
            Input_Schema, Output_Schema,
            Init_Final, Init_Hold, Init_Review,
            Epic_Final, Epic_Hold, Epic_Review, Epic_Milestone,
            Removed_Rows, Output_File_Path, Runtime_Seconds, Run_Status
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        ts, 'Prod Data Pipeline', 'prod_v2', str(input_path),
        input_schema, output_schema,
        len(init_final), len(init_hold), len(init_review),
        len(epic_final), len(epic_hold), len(epic_review), len(epic_milestone),
        removed_in + removed_ep, str(out_path), elapsed, status
    ))
    conn.commit()
    log(f"Run logged : run_history.Pipeline_Runs_Prod — Run_ID: {ts}", 1)


# ── Main ──────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Planview Prod Data Pipeline")
    parser.add_argument("--config", default=None)
    args = parser.parse_args()
    load_config(args.config)

    start = datetime.now()
    ts    = start.strftime("%Y%m%d_%H%M%S")
    log_file = init_logger(ts, OUTPUT_FOLDER)

    log(SEPARATOR)
    log("  Planview2026PlatinumConsolidatedUpdated")
    log(f"  Input  : {INPUT_FILE}")
    log(f"  Output : {OUTPUT_FOLDER}")
    log(f"  Started: {start.strftime('%Y-%m-%d %H:%M:%S')}")
    log(SEPARATOR)

    # ── Step 1 — Read input files ─────────────────────────────
    log_step("1/5", "Reading input files (3-row header)...")
    ip = Path(INPUT_FILE)
    df_in_raw, in_r1, in_r2 = read_3row_header(ip, INPUT_SHEET_INITS)
    df_ep_raw, ep_r1, ep_r2 = read_3row_header(ip, INPUT_SHEET_EPICS)
    log(f"Initiatives : {len(df_in_raw):,} rows | {len(df_in_raw.columns)} cols", 1)
    log(f"Epics       : {len(df_ep_raw):,} rows | {len(df_ep_raw.columns)} cols", 1)
    log("Row 1=old IDs, Row 2=new IDs, Row 3=display names used as column headers", 1)

    # ── Step 2 — Connect + load raw input to SQL ─────────────
    conn = connect_sql()
    stem = re.sub(r'[^A-Za-z0-9_]', '_', ip.stem).strip('_')
    input_schema, tbl_init, tbl_epic = create_input_schema_and_load(
        conn, df_in_raw, df_ep_raw, ip, ts)

    # ── Steps 2a–4c — Stored procedure ───────────────────────
    (df_inits, df_epics, df_deleted,
     removed_in, removed_ep, new_id_count,
     changes_in, changes_ep) = run_transform_sp(conn, ts, input_schema, stem)

    # ── Console summary of classification results ─────────────
    ivc = df_inits['Output_Segment'].value_counts()
    log(f"INITIATIVES ({len(df_inits):,} total):", 1)
    for k in ["init_bc","init_strat_inv","init_disc_other","init_bo"]:
        c = ivc.get(SEG[k], 0)
        if c: log(f"  {SEG[k]:<45}: {c:>6}", 2)
    log(f"  {'SUBTOTAL Final':<45}: {sum(ivc.get(SEG[k],0) for k in ['init_bc','init_strat_inv','init_disc_other','init_bo']):>6}", 2)
    for k in ["init_stopwork","init_pc"]:
        c = ivc.get(SEG[k], 0)
        if c: log(f"  {SEG[k]:<45}: {c:>6}", 2)
    c = ivc.get(SEG['init_pending'], 0) + ivc.get(SEG['review'], 0)
    if c: log(f"  {'Review/Pending':<45}: {c:>6}", 2)

    evc = df_epics['Output_Segment'].value_counts()
    log(f"\nEPICS ({len(df_epics):,} total):", 1)
    for k in ["lcm","le_strat","le_disc","bwt_epic"]:
        c = evc.get(SEG[k], 0)
        if c: log(f"  {SEG[k]:<45}: {c:>6}", 2)
    log(f"  {'SUBTOTAL Final':<45}: {sum(evc.get(SEG[k],0) for k in ['lcm','le_strat','le_disc','bwt_epic']):>6}", 2)
    c = evc.get(SEG['le_stopwork'], 0)
    if c: log(f"  {SEG['le_stopwork']:<45}: {c:>6}", 2)
    c = evc.get(SEG['epic_pending'], 0) + evc.get(SEG['review'], 0)
    if c: log(f"  {'Review/Pending':<45}: {c:>6}", 2)
    c = evc.get(SEG['milestone'], 0)
    if c: log(f"  {'Milestone/Risk (separate tab)':<45}: {c:>6}", 2)
    log(f"  NewID_Temp assigned: {new_id_count:,}", 1)

    # ── Step 5a — Save classified output to SQL ───────────────
    (output_schema,
     init_final, init_hold, init_review,
     epic_final, epic_hold, epic_review, epic_milestone) = create_output_schema_and_save(
        conn, df_inits, df_epics, df_deleted, stem, ts)

    # ── Step 5b — Log run to run_history ─────────────────────
    log_step("5b/5", "Logging run to run_history.Pipeline_Runs_Prod...")

    # ── Step 5c — Write Excel output ─────────────────────────
    log_step("5c/5", "Writing output Excel...")
    (out_path,
     init_final, init_hold, init_review,
     epic_final, epic_hold, epic_review, epic_milestone) = build_excel(
        df_inits, df_epics, df_deleted, df_in_raw, df_ep_raw,
        ts, ip, OUTPUT_FOLDER,
        removed_in, removed_ep, changes_in, changes_ep)

    elapsed = round((datetime.now() - start).total_seconds(), 1)

    log(f"\n{SEPARATOR}")
    log("  PIPELINE COMPLETE")
    log(f"  Output : {out_path}")
    log(f"  Log    : {log_file}")
    log(f"")
    log(f"  INITIATIVES ({len(df_inits):,} total \u2014 all statuses included):")
    log(f"    Final Output   : {len(init_final):,}")
    log(f"    Stop Work HOLD : {len(init_hold):,}  (inc. Purple Chip/DRIVE)")
    log(f"    Review/Pending : {len(init_review):,}  (SL/L5 stages undefined)")
    log(f"")
    log(f"  EPICS ({len(df_epics):,} total \u2014 all statuses included):")
    log(f"    Final Output   : {len(epic_final):,}  (incl. BwT Epics with NewID_Temp)")
    log(f"    Stop Work HOLD : {len(epic_hold):,}")
    log(f"    Review/Pending : {len(epic_review):,}")
    log(f"    Milestone/Risk : {len(epic_milestone):,}  (separate tab)")
    log(f"")
    log(f"  Deleted (L0/SL1)  : {removed_in} Initiative rows → Deleted_Records tab")
    log(f"  NewID_Temp        : {new_id_count:,} Epic records")
    log(f"  Runtime           : {elapsed}s")

    # Step 5b runs after Excel is saved so out_path is available
    log_run_history(conn, ts, ip, input_schema, output_schema,
                    init_final, init_hold, init_review,
                    epic_final, epic_hold, epic_review, epic_milestone,
                    out_path, elapsed, "Completed", removed_in, removed_ep)

    conn.close()
    log("SQL      : Connection closed", 1)
    log(SEPARATOR)


if __name__ == "__main__":
    main()
